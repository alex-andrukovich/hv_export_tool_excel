#!/usr/bin/env python3
# usage: -z "C:\xxx\xxx\xxx\xxx.zip" -e "C:\xxx\xxx\xxx\extracted"
import os
import zipfile
import logging
import functools
import argparse
import pandas as pd
import openpyxl
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from io import StringIO
import multiprocessing as mp
import gc
import random
from tqdm import tqdm

# Create a logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
# Create handlers
file_handler = logging.FileHandler('hv_export_tool_excel.log')  # Log to a file
stdout_handler = logging.StreamHandler()  # Log to stdout
# Create a formatter and add it to the handlers
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
stdout_handler.setFormatter(formatter)
# Add the handlers to the logger
logger.addHandler(file_handler)
logger.addHandler(stdout_handler)


def log_decorator(fn):
    @functools.wraps(fn)
    def wrapper(*args, **kwargs):
        logger.info(f'Function {fn.__name__} called with args: {args} and kwargs: {kwargs}')
        return fn(*args, **kwargs)
    return wrapper


@log_decorator
def get_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("-z", "--zippath", dest="zippath",  help="Enter full path to the export tool zip file e.g. \"c:\\out.zip")
    parser.add_argument("-e", "--extractpath", dest="extractpath", help="Enter full path where zip file is going to be extracted e.g. \"c:\\extracted_zip\\")
    arguments = parser.parse_args()
    if not arguments.zippath:
        parser.exit("Enter full path to the export tool zip file e.g. \"c:\\out.zip\"")
    elif not arguments.extractpath:
        parser.exit("Enter full path where zip file is going to be extracted e.g. \"c:\\extracted_zip\"")
    return arguments


@log_decorator
def unzip_all(zip_path, extract_path):
    archive_type="highend"
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_path)
        for root, dirs, files in os.walk(extract_path):
            for file in files:
                if "export_metadata" in file.lower():
                    archive_type = "midrange"
                logger.info(file)
                if file.lower().endswith('.zip'):
                    file_path = os.path.join(root, file)
                    new_extract_path = os.path.join(root, file[:-4])
                    os.makedirs(new_extract_path, exist_ok=True)
                    with zipfile.ZipFile(file_path, 'r') as inner_zip_ref:
                        inner_zip_ref.extractall(new_extract_path)
                    os.remove(file_path)
    return archive_type


# Example usage
@log_decorator
def list_extracted_csv_files(extract_path):
    extracted_files_list = []
    for root, dirs, files in os.walk(extract_path):
        for file in files:
            if file.lower().endswith('.csv') and "export_metadata" not in file.lower():
                file_path = os.path.join(root, file)
                logger.info(file_path)
                extracted_files_list.append(file_path)
    return extracted_files_list


@log_decorator
def add_charts(file, data_min_row):
    large = False
    short_file_name = file.split('\\')[-1]
    short_file_name = short_file_name.replace(".csv", ".xlsx")
    output_file = file.replace(".csv", ".xlsx")
    # Load the workbook and select the sheet
    wb = load_workbook(output_file)
    ws = wb['Sheet1']

    # Create a reference to the data for the chart
    # values = Reference(ws, min_col=2, min_row=2, max_col=ws.max_column, max_row=ws.max_row)
    remaining_col = ws.max_column
    current_col = 2
    where_to_add_chart = 5
    while remaining_col > 0:
        if remaining_col >= 250:
            values = Reference(ws, min_col=current_col, min_row=data_min_row, max_col=current_col + 250, max_row=ws.max_row)
            remaining_col = remaining_col - 250
            current_col = current_col + 250
        else:
            values = Reference(ws, min_col=current_col, min_row=data_min_row, max_col=current_col + remaining_col,
                               max_row=ws.max_row)
            remaining_col = remaining_col - 250
            current_col = current_col + remaining_col
        categories = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)

        # Create the chart
        chart = LineChart()
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = short_file_name.replace(".xlsx", "")
        # Set axis titles
        chart.x_axis.title = "DateTime"
        chart.y_axis.title = "Values"

        # Explicitly set axis lines to be visible
        chart.x_axis.majorTickMark = "in"
        chart.y_axis.majorTickMark = "in"
        chart.x_axis.minorTickMark = "in"
        chart.y_axis.minorTickMark = "in"

        # Set number format for axis labels
        chart.x_axis.number_format = 'dd-mmm-yyyy hh:mm'
        chart.y_axis.number_format = 'General'

        # Ensure tick labels are shown
        chart.x_axis.tickLblPos = 'nextTo'
        chart.y_axis.tickLblPos = 'nextTo'
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        # Add the chart to the sheet
        anchor = "C" + str(where_to_add_chart)
        ws.add_chart(chart, anchor)
        where_to_add_chart = where_to_add_chart + 60

        # Adjust the size of the chart
        chart.width = 60  # Set the width of the chart
        chart.height = 30  # Set the height of the chart
    # Save the workbook
    wb.save(output_file)
    # Close the workbook
    wb.close()


@log_decorator
def read_csv_convert_to_excel_midrange(file):
        df = pd.read_csv(file, delimiter=',')
        df['DateTime'] = pd.to_datetime(df['Date'] + ' ' + df['Time'])
        df.drop(columns=['Date', 'Time'], inplace=True)
        value_columns = [col for col in df.columns if col not in ['DateTime', 'ID', 'Date', 'Time']]
        # print(value_columns)
        pivot_df = df.pivot(index='DateTime', columns='ID', values=value_columns)
        del value_columns
        del df
        gc.collect()
        pivot_df.to_excel(file.replace(".csv", ".xlsx"))
        del pivot_df
        gc.collect()
        add_charts(file, 2)


@log_decorator
def read_csv_convert_to_excel_highend(file):
    with open(file, 'r') as csv:
        # Read all lines from the file
        lines = csv.readlines()
    # Drop the first 6 lines
    lines = lines[6:]
    fixed_lines = []
    for l in lines:
        index = int(l.split(',')[0].replace("\"", "").replace("No.", "0"))
        if 0 <= index < len(fixed_lines):
            split_l = l.split(',')
            split_l_without_index_and_date = split_l[2:]
            split_l_without_index_and_date_str = ",".join(split_l_without_index_and_date)
            fixed_lines[index] = fixed_lines[index].strip() + "," + split_l_without_index_and_date_str
        else:
            fixed_lines.insert(index, l)
    del lines
    gc.collect()
    if "PhyProcDetail_dat" not in file:
        data_str = "\n".join(fixed_lines)
        del fixed_lines
        gc.collect()
        data_io = StringIO(data_str)
        del data_str
        gc.collect()
        df = pd.read_csv(data_io, delimiter=',')
        df.reset_index(drop=True, inplace=True)
        df.drop(columns=['No.'], inplace=True)
        df.set_index('time', inplace=True)
        df.to_excel(file.replace(".csv", ".xlsx"))
        del df
        gc.collect()
        # add charts to the Excel file
        if "LU" not in file:
            add_charts(file, 1)
    else:
        mppk_data_list = []
        dataline = []
        for l in fixed_lines[1:]:
            splitlines = l.split(",")
            splitlines = splitlines[1:]
            if '-' not in splitlines[1]:
                for mppk_item in splitlines[1:]:
                    dataline.append(splitlines[0].replace('\"',''))
                    values = mppk_item.strip().split(";")
                    for value in values:
                        dataline.append(value)
                    mppk_data_list.append(dataline)
                    dataline = []
        df = pd.DataFrame(mppk_data_list, columns=['Date_and_time', 'Workload_Type', 'Worker', 'ID', 'Percent'])
        del fixed_lines
        gc.collect()
        df['Percent'] = pd.to_numeric(df['Percent'])
        df['Date_and_time'] = pd.to_datetime(df['Date_and_time'], format='%Y/%m/%d %H:%M')
        df.to_excel(file.replace(".csv", ".xlsx"))
        del df
        gc.collect()


def main():
    chunk_size = 1
    user_input = get_arguments()
    zip_path = user_input.zippath
    extract_path = user_input.extractpath
    archive_type = unzip_all(zip_path, extract_path)
    # print(archive_type)
    my_files = list_extracted_csv_files(extract_path)
    random.shuffle(my_files)
    if archive_type == "midrange":
        with mp.Pool() as pool:
            # list(pool.imap_unordered(read_csv_convert_to_excel_midrange, my_files, chunksize=chunk_size))
            for _ in tqdm(pool.imap_unordered(read_csv_convert_to_excel_midrange, my_files, chunksize=chunk_size),
                          total=len(my_files)):
                pass
    elif archive_type == "highend":
        with mp.Pool() as pool:
            # list(pool.imap_unordered(read_csv_convert_to_excel_highend, my_files, chunksize=chunk_size))
            for _ in tqdm(pool.imap_unordered(read_csv_convert_to_excel_highend, my_files, chunksize=chunk_size),
                          total=len(my_files)):
                pass


if __name__ == "__main__":
    mp.freeze_support()
    main()
